# -*- coding: utf-8 -*-
"""
Created on Wed Mar 23 19:11:06 2016

@author: Emanuele
"""

#! python3
# addImagesExcel.py - Add images to the database and write back the returned ID

import openpyxl
import requests

url = "http://replica.dhlabdemo.org:5009/api/v1/database"

print('Opening workbook')
wb = openpyxl.load_workbook('lastchargeTables.xlsx')
sheets = ['Table46', 'Table2', 'Table25', 'Table45']

for s in range(0, len(sheets)):
    print('Reading', sheets[s])
    sheet = wb.get_sheet_by_name(sheets[s])
    print('Reading rows')
    for r in range(2, sheet.max_row+1): # skip the first row
		
        imageURL = sheet['A' + str(r)].value        
        if(imageURL is None):
            print("Missing image URL at ", sheet, "-", r)
            exit()       
        author = sheet['B' + str(r)].value
        if(author is None):
            author = ""
        title = sheet['C' + str(r)].value
        if(title is None):
            title = ""
        date = sheet['D' + str(r)].value
        if(date is None):
            date = ""
        school = sheet['E' + str(r)].value
        if(school is None):
            school = ""
        form = sheet['F' + str(r)].value
        if(form is None):
            form = ""
        typeImg = sheet['G' + str(r)].value
        if(typeImg is None):
            typeImg = ""
        origin = sheet['H' + str(r)].value
        if(origin is None):
            origin = ""
        tableNumber = sheet['I' + str(r)].value
        if(origin is None):
            origin = ""
        pictureNumber = sheet['J' + str(r)].value
        if(origin is None):
            origin = ""
        webpageURL = sheet['K' + str(r)].value
        if(webpageURL is None):
            webpageURL = ""
        
        
        jsonData = {"image_url": str(imageURL),
                "metadata": {"author": str(author),
                             "title": str(title),
                             "date": str(date),
                             "school": str(school),
                             "form": str(form),
                             "type": str(typeImg),
                             "table of the atlas": str(tableNumber),
                             "number of the picture": str(pictureNumber)
                             },
               "origin": str(origin),
               "webpage_url": str(webpageURL)
               }
        req = requests.post(url, json = jsonData)
        print(req)

        ID = str(req.json())
        
        #update worksheet
        sheet.cell(row=r, column=12).value = ID #first index is 1. L <-> 12
        
wb.save('2004Tables.xlsx')
 
        

    

