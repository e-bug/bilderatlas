# -*- coding: utf-8 -*-
"""
Created on Wed Mar 23 18:40:06 2016

@author: e-bug
"""

#! python3
# searchImagesExcel.py - Search images in the database and write back the metadate of the results; 
#                        download the images of the results ordering them by their score

import openpyxl
import requests
import json
import time
import os
import urllib.request

# replace any invalid character in a filename with '_'
def valid_filename(filename):
    invalid_chars = "#<$+%>!`&*‘'|{?“=}/:\@"
    l = list(filename)
    for i in range(0,len(l)):
        if l[i] in invalid_chars:
            l[i]='_'
    return ''.join(l)
    

url = "http://replica.dhlabdemo.org:5009/api/v1/search/ids"
nResults = 30	#number of results returned by the database
excel_file = 'searchIDs.xlsx'
sheets = ['Query']


wb = openpyxl.load_workbook(excel_file)
sheet = wb.get_sheet_by_name(sheets[0])

# create directory for the current query having as name the current time
dirName = time.strftime("%Y%m%d%H%M%S")
if not os.path.exists(dirName):
    os.makedirs(dirName)

# collect positive and negative IDs
posIDs = []
negIDs = []
for r in range(2, sheet.max_row+1): # skip the first row
    posIDs.append(sheet['A' + str(r)].value)
    negIDs.append(sheet['B' + str(r)].value)
    
posIDs = [v for v in posIDs if v is not None]
negIDs = [v for v in negIDs if v is not None]

jsonData = {"positive_image_ids": posIDs,
            "negative_image_ids": negIDs,
            "nb_results": nResults
            }

# send HTTP POST 
req = requests.post(url, json = jsonData)
print(req)

# collect the result
response = json.dumps(req.json())
jsonObject = json.loads(response)
results = jsonObject['results']

for i in range(0, nResults): 
    # retrieve metadata associated to the results
    sheet.cell(row=i+2, column=4).value = abs(results[i]['score'])
    sheet.cell(row=i+2, column=5).value = results[i]['id']
    sheet.cell(row=i+2, column=6).value = results[i]['metadata']['author']
    sheet.cell(row=i+2, column=7).value = results[i]['metadata']['title']
    sheet.cell(row=i+2, column=8).value = results[i]['metadata']['date']
    if 'table of the atlas' in results[i]['metadata']: # optional field, used only in the bilderatlas context
        sheet.cell(row=i+2, column=9).value = results[i]['metadata']['table of the atlas']
    sheet.cell(row=i+2, column=10).value = results[i]['image_url']
    if 'webpage_url' in results[i]: # optional field
        sheet.cell(row=i+2, column=11).value = results[i]['webpage_url']
    
    # download image
    imageName = dirName+'/'+str(i)+'_'+valid_filename(results[i]['metadata']['title'])+'.jpg'
    urllib.request.urlretrieve(results[i]['image_url'], imageName)

    
wb.save(dirName +'/'+dirName+'.xlsx')



    

